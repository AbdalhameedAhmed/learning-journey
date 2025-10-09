import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


async def generate_report(headers, report_data, title):
    wb = Workbook()
    ws = wb.active
    if not ws:
        return None

    ws.title = title

    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    for row_data in report_data:
        ws.append(row_data)

    # 4. Save the workbook to a BytesIO object (in-memory file)
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)  # Rewind the stream to the beginning

    return excel_stream
